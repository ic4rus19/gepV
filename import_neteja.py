from datetime import datetime, date
from openpyxl import load_workbook

from app import app, db, NetejaSetmanal  # usa tu app.py (sin paquete)


EXCEL_PATH = "data/neteja_setmanal.xlsx"


def parse_date(value):
    """Convierte fecha Excel/str a datetime.date o None."""
    if value is None or value == "":
        return None

    if isinstance(value, date):
        return value  # ya es date/datetime

    if isinstance(value, str):
        s = value.strip()
        # formatos habituales: 31/01/2026 o 2026-01-31
        for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
            try:
                return datetime.strptime(s, fmt).date()
            except ValueError:
                pass

    return None


def parse_time(value):
    """Guarda hora como texto HH:MM (robusto en SQLite)."""
    if value is None:
        return None

    if isinstance(value, str):
        s = value.strip()
        return s if s else None

    # Si excel trae datetime/time, openpyxl a veces da datetime
    if hasattr(value, "hour") and hasattr(value, "minute"):
        return f"{value.hour:02d}:{value.minute:02d}"

    return str(value)


def main():
    wb = load_workbook(EXCEL_PATH)
    ws = wb.active

    # Leer cabeceras (primera fila)
    headers = []
    for cell in ws[1]:
        headers.append((str(cell.value).strip().lower() if cell.value else ""))

    required = ["nom", "torn", "dia", "centre", "inici", "fi", "temps"]

    missing = [c for c in required if c not in headers]
    if missing:
        raise ValueError(f"Faltan columnas en el Excel: {missing}. Cabeceras detectadas: {headers}")

    idx = {h: headers.index(h) for h in required}

    rows_to_insert = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        nom = r[idx["nom"]]
        torn = r[idx["torn"]]
        dia = parse_date(r[idx["dia"]])
        centre = r[idx["centre"]]
        inici = parse_time(r[idx["inici"]])
        fi = parse_time(r[idx["fi"]])
        temps = r[idx["temps"]]

        rows_to_insert.append(
            NetejaSetmanal(
                nom=str(nom).strip() if nom else None,
                torn=str(torn).strip().upper() if torn else None,
                dia=dia,
                centre=str(centre).strip() if centre else None,
                inici=inici,
                fi=fi,
                temps=str(temps).strip() if temps else None,
            )
        )

    with app.app_context():
        # borrar y recargar (simple y fiable)
        NetejaSetmanal.query.delete()
        db.session.commit()

        db.session.bulk_save_objects(rows_to_insert)
        db.session.commit()

        print(f"✅ Import OK: {len(rows_to_insert)} filas a neteja_setmanal")


if __name__ == "__main__":
    main()
