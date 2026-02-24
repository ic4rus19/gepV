# ------------------------------------------------------------
# Importador de claus con control de cambios por HASH
# ------------------------------------------------------------

import os
import hashlib
import warnings
from openpyxl import load_workbook
from app import app, db, Claus

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

EXCEL_PATH = "data/claus.xlsx"
HASH_PATH = "data/.claus_hash"


def calcular_hash(path):
    sha256 = hashlib.sha256()
    with open(path, "rb") as f:
        for bloque in iter(lambda: f.read(4096), b""):
            sha256.update(bloque)
    return sha256.hexdigest()


def main():

    if not os.path.exists(EXCEL_PATH):
        print(f"❌ No se encuentra el archivo: {EXCEL_PATH}")
        return

    nuevo_hash = calcular_hash(EXCEL_PATH)

    # Comprobar si ya existe hash anterior
    if os.path.exists(HASH_PATH):
        with open(HASH_PATH, "r") as f:
            hash_guardado = f.read().strip()

        if nuevo_hash == hash_guardado:
            print("ℹ️ El fichero no ha cambiado. No se realiza importación.")
            return

    print("📂 Archivo modificado. Iniciando importación...")

    wb = load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    ws = wb.active

    headers = [(str(c.value).strip().lower() if c.value else "") for c in ws[1]]

    required = ["nom_porta", "nom_espai", "armari", "num", "actuals"]
    missing = [c for c in required if c not in headers]

    if missing:
        print(f"❌ Faltan columnas: {missing}")
        return

    idx = {h: headers.index(h) for h in required}
    rows_to_insert = []

    for r in ws.iter_rows(min_row=2, values_only=True):

        nom_porta = r[idx["nom_porta"]]
        nom_espai = r[idx["nom_espai"]]
        armari = r[idx["armari"]]
        num = r[idx["num"]]
        actuals = r[idx["actuals"]]

        if nom_porta is None and nom_espai is None and armari is None and num is None:
            continue

        nom_porta = str(nom_porta).strip() if nom_porta else None
        nom_espai = str(nom_espai).strip() if nom_espai else None
        armari = str(armari).strip() if armari else None

        try:
            num = int(num) if num is not None else 0
        except Exception:
            num = 0

        try:
            actuals = int(actuals) if actuals is not None else None
        except Exception:
            actuals = None

        if not nom_porta or not nom_espai or not armari:
            continue

        rows_to_insert.append(
            Claus(
                nom_porta=nom_porta,
                nom_espai=nom_espai,
                armari=armari,
                num=num,
                actuals=actuals
            )
        )

    with app.app_context():

        db.session.query(Claus).delete()
        db.session.commit()

        db.session.bulk_save_objects(rows_to_insert)
        db.session.commit()

        print(f"✅ Import OK: {len(rows_to_insert)} filas a claus")

    # Guardar nuevo hash
    with open(HASH_PATH, "w") as f:
        f.write(nuevo_hash)


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print(f"❌ Error durante la importación: {e}")